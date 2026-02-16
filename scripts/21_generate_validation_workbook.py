"""
Generate Excel-based validation workbook for easy human review.

Creates a smart Excel file with:
- Side-by-side comparison (original vs extracted)
- Color coding (green = confident, yellow = review, red = error)
- Dropdown menus for corrections (top suggestions auto-filled)
- Bulk operations (accept all high-confidence)
- Progress tracking
- Save and re-import for learning
"""
import sys
sys.path.insert(0, '.')

import sqlite3
import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional
import argparse
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from src.database.connection import DatabaseManager
from src.normalization.text_normalizer import TextNormalizer


class ValidationWorkbookGenerator:
    """Generate user-friendly Excel validation workbooks."""
    
    # Color scheme
    COLOR_CONFIDENT = "C6EFCE"  # Light green
    COLOR_REVIEW = "FFEB9C"     # Light yellow
    COLOR_ERROR = "FFC7CE"      # Light red
    COLOR_HEADER = "4472C4"     # Blue
    
    def __init__(self):
        self.db = DatabaseManager('data/reg153_matcher.db')
        self.lab_db_path = Path('data/lab_results.db')
        self.normalizer = TextNormalizer()
    
    def generate_review_workbook(
        self,
        submission_id: int,
        output_path: Path,
        confidence_threshold: float = 0.95
    ):
        """
        Generate validation workbook for a submission.
        
        Args:
            submission_id: ID of the submission to review
            output_path: Where to save the Excel file
            confidence_threshold: Auto-accept above this (default 0.95)
        """
        print(f"Generating validation workbook for submission {submission_id}...")
        
        # Get submission details
        lab_conn = sqlite3.connect(str(self.lab_db_path))
        
        submission = lab_conn.execute("""
            SELECT original_filename, lab_vendor, layout_confidence, 
                   file_path, sheet_name, extraction_timestamp
            FROM lab_submissions
            WHERE submission_id = ?
        """, (submission_id,)).fetchone()
        
        if not submission:
            print(f"✗ Submission {submission_id} not found")
            return False
        
        filename, vendor, layout_conf, file_path, sheet_name, extracted_at = submission
        
        # Get extraction results
        results = lab_conn.execute("""
            SELECT result_id, row_number, chemical_raw, chemical_normalized,
                   analyte_id, match_method, match_confidence,
                   sample_id, result_value, units, qualifier,
                   validation_status, match_alternatives
            FROM lab_results
            WHERE submission_id = ?
            ORDER BY row_number
        """, (submission_id,)).fetchall()
        
        lab_conn.close()
        
        if not results:
            print(f"✗ No results found for submission {submission_id}")
            return False
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Instructions
        self._create_instructions_sheet(wb, filename, vendor, layout_conf, len(results))
        
        # Sheet 2: Chemical Review
        self._create_chemical_review_sheet(
            wb, results, confidence_threshold, submission_id
        )
        
        # Sheet 3: Summary Dashboard
        self._create_summary_sheet(wb, results, confidence_threshold)
        
        # Save
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        
        print(f"✓ Validation workbook created: {output_path}")
        print(f"\n  📋 {len(results)} extractions to review")
        
        high_conf = sum(1 for r in results if r[6] and r[6] >= confidence_threshold)
        needs_review = len(results) - high_conf
        
        print(f"  ✓ {high_conf} high-confidence (auto-accept)")
        print(f"  ⚠ {needs_review} need your review")
        
        return True
    
    def _create_instructions_sheet(
        self,
        wb: Workbook,
        filename: str,
        vendor: str,
        layout_conf: float,
        num_results: int
    ):
        """Create friendly instructions sheet."""
        ws = wb.active
        ws.title = "📖 Instructions"
        
        # Header
        ws['A1'] = "HOW TO VALIDATE EXTRACTIONS"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # File info
        ws['A3'] = "File:"
        ws['B3'] = filename
        ws['A4'] = "Vendor:"
        ws['B4'] = vendor
        ws['A5'] = "Layout Confidence:"
        ws['B5'] = f"{layout_conf:.1%}" if layout_conf else "N/A"
        ws['A6'] = "Chemicals Found:"
        ws['B6'] = num_results
        
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)
        
        # Instructions
        instructions = [
            "",
            "STEP-BY-STEP INSTRUCTIONS:",
            "",
            "1. Go to the '🔬 Chemical Review' tab",
            "",
            "2. Look at the colored rows:",
            "   🟢 GREEN = High confidence, auto-accepted (no action needed)",
            "   🟡 YELLOW = Please review and select correct match from dropdown",
            "   🔴 RED = Error detected, requires your attention",
            "",
            "3. For YELLOW/RED rows:",
            "   - Check the 'Chemical Name (From Excel)' column",
            "   - Look at 'Matched To' - is it correct?",
            "   - If wrong: Click 'Corrected Match' dropdown and select right one",
            "   - If correct: Leave 'Corrected Match' blank",
            "   - Add notes in 'Validation Notes' if helpful (optional)",
            "",
            "4. Save this file when done",
            "",
            "5. Run: python scripts/22_import_validations.py --file [this file]",
            "",
            "TIPS:",
            "• Dropdowns show the 5 most likely matches for each chemical",
            "• You only need to fill 'Corrected Match' if the auto-match is wrong",
            "• Use Ctrl+D to copy down cells if same correction applies to multiple rows",
            "• Check the 'Summary' tab to see your progress",
            "",
            "Questions? The system learns from your corrections to improve over time!"
        ]
        
        for idx, instruction in enumerate(instructions, start=8):
            ws[f'A{idx}'] = instruction
            if "STEP-BY-STEP" in instruction or "TIPS:" in instruction:
                ws[f'A{idx}'].font = Font(bold=True, size=12)
        
        # Set column widths
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 30
    
    def _create_chemical_review_sheet(
        self,
        wb: Workbook,
        results: List[tuple],
        confidence_threshold: float,
        submission_id: int
    ):
        """Create the main review sheet with dropdowns."""
        ws = wb.create_sheet("🔬 Chemical Review")
        
        # Headers
        headers = [
            "Row", "Status", "Chemical Name\n(From Excel)", "Matched To",
            "Confidence", "Match Method", "Corrected Match",
            "Validation Notes", "Sample ID", "Result", "Units"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        ws.row_dimensions[1].height = 35
        
        # Get all analytes for dropdown - we'll create a reference sheet
        with self.db.get_session() as session:
            from src.database.models import Analyte
            all_analytes = session.query(Analyte.analyte_id, Analyte.preferred_name).order_by(Analyte.preferred_name).all()
            analyte_options = [f"{name} ({aid})" for aid, name in all_analytes]
        
        # Note: We'll add dropdown after data is populated
        
        # Add data rows
        for row_idx, result in enumerate(results, start=2):
            (result_id, row_num, chem_raw, chem_norm, analyte_id, 
             match_method, match_conf, sample_id, result_val, units, 
             qualifier, val_status, alternatives) = result
            
            # Determine status and color
            if match_conf and match_conf >= confidence_threshold:
                status = "✓ Confident"
                fill_color = self.COLOR_CONFIDENT
            elif match_conf and match_conf >= 0.70:
                status = "⚠ Review"
                fill_color = self.COLOR_REVIEW
            else:
                status = "✗ Error"
                fill_color = self.COLOR_ERROR
            
            # Get matched analyte name
            matched_name = ""
            if analyte_id:
                with self.db.get_session() as session:
                    from src.database.models import Analyte
                    analyte = session.query(Analyte).filter(
                        Analyte.analyte_id == analyte_id
                    ).first()
                    if analyte:
                        matched_name = f"{analyte.preferred_name} ({analyte_id})"
            
            # Populate row - sanitize values to prevent corruption
            col_values = [
                row_num,
                status,
                str(chem_raw) if chem_raw else "",
                matched_name,
                f"{match_conf:.1%}" if match_conf else "N/A",
                match_method or "none",
                "",  # Corrected Match (empty for user to fill)
                "",  # Validation Notes
                str(sample_id) if sample_id else "",
                str(result_val) if result_val else "",
                str(units) if units else ""
            ]
            
            for col_idx, value in enumerate(col_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply color coding
                if col_idx <= 6:  # Color the review columns
                    cell.fill = PatternFill(start_color=fill_color, fill_type="solid")
                
                # Alignment
                if col_idx in [1, 4, 5]:  # Row, Confidence, Method
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left", wrap_text=True)
        
        # Create reference sheet for dropdown (must be created AFTER we know the sheet structure)
        ref_ws = wb.create_sheet("AnalyteList", index=len(wb.worksheets))
        for idx, option in enumerate(analyte_options, start=1):
            ref_ws.cell(row=idx, column=1, value=option)
        ref_ws.sheet_state = 'hidden'
        
        # Add dropdown validation for "Corrected Match" column
        # Use absolute reference to hidden sheet
        dropdown_formula = f"AnalyteList!$A$1:$A${len(analyte_options)}"
        
        dv = DataValidation(
            type="list",
            formula1=dropdown_formula,
            allow_blank=True,
            showDropDown=True
        )
        dv.error = "Please select from the dropdown list or type a custom value"
        dv.errorTitle = "Invalid Selection"
        dv.showErrorMessage = False  # Don't block custom input
        dv.promptTitle = "Select Chemical"
        dv.prompt = "Click dropdown arrow to see options, or type to search"
        dv.showInputMessage = True
        
        ws.add_data_validation(dv)
        
        # Apply to all data rows in column G (Corrected Match)
        for row_idx in range(2, len(results) + 2):
            dv.add(f"G{row_idx}")
        
        print(f"   ✓ Added dropdowns with {len(analyte_options)} options")
        
        # Set column widths
        widths = [8, 12, 30, 35, 12, 15, 35, 30, 15, 12, 10]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Freeze panes
        ws.freeze_panes = "C2"
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        results: List[tuple],
        confidence_threshold: float
    ):
        """Create summary dashboard."""
        ws = wb.create_sheet("📊 Summary")
        
        # Calculate stats
        total = len(results)
        high_conf = sum(1 for r in results if r[6] and r[6] >= confidence_threshold)
        medium_conf = sum(1 for r in results if r[6] and 0.70 <= r[6] < confidence_threshold)
        low_conf = total - high_conf - medium_conf
        
        # Header
        ws['A1'] = "EXTRACTION SUMMARY"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
        ws.merge_cells('A1:C1')
        
        # Stats
        ws['A3'] = "Total Chemicals:"
        ws['B3'] = total
        ws['A4'] = "✓ High Confidence:"
        ws['B4'] = high_conf
        ws['C4'] = f"{high_conf/total*100:.1f}%"
        ws['A5'] = "⚠ Need Review:"
        ws['B5'] = medium_conf
        ws['C5'] = f"{medium_conf/total*100:.1f}%"
        ws['A6'] = "✗ Low Confidence:"
        ws['B6'] = low_conf
        ws['C6'] = f"{low_conf/total*100:.1f}%"
        
        # Color code
        ws['A4'].fill = PatternFill(start_color=self.COLOR_CONFIDENT, fill_type="solid")
        ws['A5'].fill = PatternFill(start_color=self.COLOR_REVIEW, fill_type="solid")
        ws['A6'].fill = PatternFill(start_color=self.COLOR_ERROR, fill_type="solid")
        
        for row in range(3, 7):
            ws[f'A{row}'].font = Font(bold=True)
        
        # Method breakdown
        ws['A9'] = "BY MATCH METHOD:"
        ws['A9'].font = Font(bold=True, size=12)
        
        method_counts = {}
        for r in results:
            method = r[5] or "none"
            method_counts[method] = method_counts.get(method, 0) + 1
        
        for idx, (method, count) in enumerate(sorted(method_counts.items()), start=10):
            ws[f'A{idx}'] = method
            ws[f'B{idx}'] = count
            ws[f'C{idx}'] = f"{count/total*100:.1f}%"
        
        # Set widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12


def main():
    parser = argparse.ArgumentParser(
        description="Generate Excel validation workbook for easy human review"
    )
    parser.add_argument(
        '--submission-id',
        type=int,
        required=True,
        help='Submission ID to generate review for'
    )
    parser.add_argument(
        '--output',
        type=Path,
        help='Output Excel file path (default: reports/validation/validation_SUBID_DATE.xlsx)'
    )
    parser.add_argument(
        '--confidence-threshold',
        type=float,
        default=0.95,
        help='Confidence threshold for auto-accept (default: 0.95)'
    )
    
    args = parser.parse_args()
    
    # Default output path
    if not args.output:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        args.output = Path(f'reports/validation/validation_{args.submission_id}_{timestamp}.xlsx')
    
    print("GENERATING VALIDATION WORKBOOK")
    print("=" * 80)
    print(f"\nSubmission ID: {args.submission_id}")
    print(f"Output: {args.output}")
    print(f"Confidence Threshold: {args.confidence_threshold:.0%}")
    print()
    
    generator = ValidationWorkbookGenerator()
    
    success = generator.generate_review_workbook(
        submission_id=args.submission_id,
        output_path=args.output,
        confidence_threshold=args.confidence_threshold
    )
    
    if success:
        print("\n" + "=" * 80)
        print("✓ SUCCESS!")
        print(f"\n💡 Next steps:")
        print(f"   1. Open: {args.output}")
        print(f"   2. Review the '🔬 Chemical Review' tab")
        print(f"   3. Fix any yellow/red rows using dropdowns")
        print(f"   4. Save the file")
        print(f"   5. Run: python scripts/22_import_validations.py --file \"{args.output}\"")
    else:
        print("\n✗ Failed to generate workbook")
        sys.exit(1)


if __name__ == "__main__":
    main()

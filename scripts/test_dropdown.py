"""
Test if Excel dropdowns are working in openpyxl.
Creates a simple test workbook with dropdowns to verify functionality.
"""
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

# Create simple workbook
wb = Workbook()
ws = wb.active
ws.title = "Test"

# Add some data
ws['A1'] = "Chemical"
ws['B1'] = "Selection"

ws['A2'] = "Benzene"
ws['A3'] = "Toluene"
ws['A4'] = "Xylene"

# Create reference sheet
ref = wb.create_sheet("Options")
options = ["Option 1", "Option 2", "Option 3", "Option 4", "Option 5"]
for idx, opt in enumerate(options, start=1):
    ref.cell(row=idx, column=1, value=opt)
ref.sheet_state = 'hidden'

# Add dropdown to column B
dv = DataValidation(
    type="list",
    formula1="Options!$A$1:$A$5",
    allow_blank=True,
    showDropDown=True
)
dv.prompt = "Select an option"
dv.promptTitle = "Choose"
dv.showInputMessage = True

ws.add_data_validation(dv)
dv.add("B2:B4")

# Save
output = Path("reports/validation/dropdown_test.xlsx")
output.parent.mkdir(exist_ok=True)
wb.save(str(output))

print(f"✓ Created test file: {output}")
print(f"  Open it and check if column B has dropdowns")
print(f"  If YES: openpyxl dropdowns work!")
print(f"  If NO: Excel version issue or openpyxl bug")

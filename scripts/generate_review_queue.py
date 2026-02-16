"""
Smart review queue generator with variant clustering.

Creates Excel review queues from recent unknowns and low-confidence matches,
with intelligent clustering of similar variants.

Usage:
    python scripts/generate_review_queue.py --output reports/review_queue.xlsx
    python scripts/generate_review_queue.py --days 7 --min-frequency 2
"""

import argparse
import json
import logging
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.database.connection import DatabaseManager
from src.database.models import MatchDecision, Analyte, LabVariant
from src.database import crud_new
from src.learning.variant_clustering import VariantClusterer
from src.matching.resolution_engine import ResolutionEngine
from sqlalchemy import select, func, and_

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/review_queue_generation.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def collect_unknowns_and_low_confidence(
    session: Session,
    days: int = 30,
    confidence_threshold: float = 0.85,
    min_frequency: int = 1
) -> List[Dict]:
    """
    Collect unknown and low-confidence matches from recent decisions.
    
    Args:
        session: Database session
        days: Number of days to look back
        confidence_threshold: Include matches below this confidence
        min_frequency: Minimum frequency to include
    
    Returns:
        List of variant dictionaries with metadata
    """
    logger.info(f"Collecting unknowns and low-confidence matches from last {days} days")
    
    cutoff = datetime.utcnow() - timedelta(days=days)
    
    # Query match decisions
    stmt = select(MatchDecision).where(
        MatchDecision.decision_timestamp >= cutoff,
        # Either unknown or low confidence
        (MatchDecision.matched_analyte_id.is_(None)) |
        (MatchDecision.confidence_score < confidence_threshold)
    )
    
    decisions = session.execute(stmt).scalars().all()
    
    logger.info(f"Found {len(decisions)} candidate decisions")
    
    # Group by input_text and count frequency
    variant_data = {}
    
    for decision in decisions:
        text = decision.input_text.strip()
        
        if text not in variant_data:
            variant_data[text] = {
                'raw_variant': text,
                'frequency': 0,
                'best_match_id': None,
                'best_match_name': None,
                'best_confidence': 0.0,
                'lab_vendor': None,
                'matrix': None,
                'top_candidates': []
            }
        
        variant_data[text]['frequency'] += 1
        
        # Track best match seen
        if decision.confidence_score > variant_data[text]['best_confidence']:
            variant_data[text]['best_match_id'] = decision.matched_analyte_id
            variant_data[text]['best_confidence'] = decision.confidence_score
            
            if decision.matched_analyte_id:
                analyte = crud_new.get_analyte_by_id(session, decision.matched_analyte_id)
                if analyte:
                    variant_data[text]['best_match_name'] = analyte.preferred_name
        
        # Store top candidates
        if decision.top_k_candidates:
            variant_data[text]['top_candidates'] = decision.top_k_candidates
    
    # Filter by minimum frequency
    variants = [
        v for v in variant_data.values()
        if v['frequency'] >= min_frequency
    ]
    
    logger.info(f"Collected {len(variants)} unique variants (min frequency: {min_frequency})")
    
    # Try to enrich with lab variant data
    for variant in variants:
        lab_variant = session.execute(
            select(LabVariant).where(
                LabVariant.observed_text == variant['raw_variant']
            ).limit(1)
        ).scalar_one_or_none()
        
        if lab_variant:
            variant['lab_vendor'] = lab_variant.lab_vendor
            variant['matrix'] = lab_variant.matrix
            variant['frequency'] = max(variant['frequency'], lab_variant.frequency_count)
    
    # Sort by frequency (descending)
    variants.sort(key=lambda x: x['frequency'], reverse=True)
    
    return variants


def cluster_variants(variants: List[Dict]) -> List[Dict]:
    """
    Cluster similar variants together.
    
    Args:
        variants: List of variant dictionaries
    
    Returns:
        List of variants with cluster_id added
    """
    if not variants:
        return []
    
    logger.info(f"Clustering {len(variants)} variants")
    
    clusterer = VariantClusterer(similarity_threshold=0.85)
    
    # Extract just the text
    variant_texts = [v['raw_variant'] for v in variants]
    
    # Run clustering
    clusters = clusterer.cluster_similar_unknowns(variant_texts)
    
    logger.info(f"Found {len(clusters)} clusters")
    
    # Build text->cluster mapping
    text_to_cluster = {}
    for cluster_id, cluster in enumerate(clusters):
        anchor = cluster['anchor']
        text_to_cluster[anchor] = cluster_id
        
        for similar_text, score in cluster['similar_variants']:
            text_to_cluster[similar_text] = cluster_id
    
    # Add cluster IDs to variants
    for variant in variants:
        text = variant['raw_variant']
        variant['cluster_id'] = text_to_cluster.get(text, -1)
    
    # Sort by cluster_id, then frequency
    variants.sort(key=lambda x: (x['cluster_id'], -x['frequency']))
    
    return variants


def generate_top_suggestions(
    variant_text: str,
    existing_candidates: List[Dict],
    session: Session,
    top_k: int = 3
) -> List[Dict]:
    """
    Generate top match suggestions for a variant.
    
    Args:
        variant_text: Variant text
        existing_candidates: Candidates from match decision
        session: Database session
        top_k: Number of suggestions
    
    Returns:
        List of suggestion dicts with analyte_id, name, confidence
    """
    # If we have existing candidates, use those
    if existing_candidates and len(existing_candidates) > 0:
        suggestions = []
        for candidate in existing_candidates[:top_k]:
            analyte_id = candidate.get('analyte_id')
            if analyte_id:
                analyte = crud_new.get_analyte_by_id(session, analyte_id)
                if analyte:
                    suggestions.append({
                        'analyte_id': analyte_id,
                        'name': analyte.preferred_name,
                        'confidence': candidate.get('confidence', 0.0)
                    })
        
        if len(suggestions) >= top_k:
            return suggestions[:top_k]
    
    # Otherwise, run fresh matching
    try:
        engine = ResolutionEngine(session)
        result = engine.resolve(variant_text, confidence_threshold=0.5)
        
        suggestions = []
        for candidate in result.all_candidates[:top_k]:
            suggestions.append({
                'analyte_id': candidate.analyte_id,
                'name': candidate.preferred_name,
                'confidence': candidate.confidence
            })
        
        return suggestions
        
    except Exception as e:
        logger.warning(f"Failed to generate suggestions for '{variant_text}': {e}")
        return []


def create_excel_review_queue(
    variants: List[Dict],
    output_path: str,
    session: Session
) -> None:
    """
    Create Excel file with validation dropdowns and formatting.
    
    Args:
        variants: List of variant dictionaries
        output_path: Path to output Excel file
        session: Database session
    """
    logger.info(f"Creating Excel review queue with {len(variants)} variants")
    
    # Prepare data rows
    rows = []
    
    for variant in variants:
        # Generate suggestions
        suggestions = generate_top_suggestions(
            variant['raw_variant'],
            variant.get('top_candidates', []),
            session,
            top_k=3
        )
        
        # Pad suggestions to 3
        while len(suggestions) < 3:
            suggestions.append({'name': '', 'confidence': 0.0})
        
        row = {
            'raw_variant': variant['raw_variant'],
            'frequency': variant['frequency'],
            'lab_vendor': variant.get('lab_vendor', ''),
            'matrix': variant.get('matrix', ''),
            'cluster_id': variant.get('cluster_id', -1),
            'suggested_match_1': suggestions[0]['name'],
            'confidence_1': round(suggestions[0]['confidence'], 3),
            'suggested_match_2': suggestions[1]['name'],
            'confidence_2': round(suggestions[1]['confidence'], 3),
            'suggested_match_3': suggestions[2]['name'],
            'confidence_3': round(suggestions[2]['confidence'], 3),
            'chosen_match': '',  # To be filled by human
            'validation_confidence': '',  # To be filled by human
            'notes': ''
        }
        
        rows.append(row)
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Get all analyte names for dropdown
    all_analytes = session.execute(select(Analyte.preferred_name)).scalars().all()
    analyte_names = sorted(set(all_analytes))
    
    logger.info(f"Loaded {len(analyte_names)} analyte names for dropdown")
    
    # Write to Excel with formatting
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Queue"
    
    # Write headers
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Color-code clusters
            if headers[col_idx-1] == 'cluster_id' and value >= 0:
                # Alternate cluster colors
                colors = ["FFE5E5", "E5F5FF", "FFF5E5", "E5FFE5"]
                color = colors[value % len(colors)]
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # Add data validation for chosen_match column (dropdown)
    chosen_match_col_idx = headers.index('chosen_match') + 1
    validation_conf_col_idx = headers.index('validation_confidence') + 1
    
    # Create analyte dropdown (limited to 255 chars per dropdown item due to Excel limitations)
    # For large lists, we'll create a reference sheet
    if len(', '.join(analyte_names[:100])) > 255:
        # Create reference sheet with all analyte names
        ref_ws = wb.create_sheet("Analyte Names")
        for idx, name in enumerate(analyte_names, start=1):
            ref_ws.cell(row=idx, column=1, value=name)
        
        # Create data validation referencing the sheet
        dv = DataValidation(
            type="list",
            formula1=f"='Analyte Names'!$A$1:$A${len(analyte_names)}",
            allow_blank=True
        )
        dv.errorTitle = "Invalid Entry"
        dv.error = "Please select from the dropdown list"
        
    else:
        # Simple dropdown
        dropdown_list = ', '.join(analyte_names + ["UNKNOWN"])
        dv = DataValidation(
            type="list",
            formula1=f'"{dropdown_list}"',
            allow_blank=True
        )
    
    dv.add(f"{chr(64 + chosen_match_col_idx)}2:{chr(64 + chosen_match_col_idx)}{len(rows) + 1}")
    ws.add_data_validation(dv)
    
    # Add validation confidence dropdown
    conf_dv = DataValidation(
        type="list",
        formula1='"HIGH,MEDIUM,LOW"',
        allow_blank=True
    )
    conf_dv.add(f"{chr(64 + validation_conf_col_idx)}2:{chr(64 + validation_conf_col_idx)}{len(rows) + 1}")
    ws.add_data_validation(conf_dv)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    wb.save(output_path)
    
    logger.info(f"Excel review queue written to: {output_path}")


def main():
    """Main entry point for review queue generation."""
    parser = argparse.ArgumentParser(
        description="Generate smart review queue with variant clustering",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Generate review queue from last 30 days
  python scripts/generate_review_queue.py --output reports/review_queue.xlsx
  
  # Last 7 days with minimum frequency of 2
  python scripts/generate_review_queue.py --days 7 --min-frequency 2 --output queue.xlsx
  
  # Include low-confidence matches below 0.80
  python scripts/generate_review_queue.py --confidence-threshold 0.80
        """
    )
    
    parser.add_argument('--output', '-o', required=True, 
                        help='Output Excel file path')
    parser.add_argument('--days', '-d', type=int, default=30,
                        help='Number of days to look back (default: 30)')
    parser.add_argument('--confidence-threshold', '-t', type=float, default=0.85,
                        help='Include matches below this confidence (default: 0.85)')
    parser.add_argument('--min-frequency', '-f', type=int, default=1,
                        help='Minimum frequency to include variant (default: 1)')
    parser.add_argument('--no-clustering', action='store_true',
                        help='Disable variant clustering')
    parser.add_argument('--database', help='Path to database file')
    
    args = parser.parse_args()
    
    try:
        logger.info("Starting review queue generation")
        logger.info(f"Parameters: days={args.days}, confidence_threshold={args.confidence_threshold}, "
                   f"min_frequency={args.min_frequency}")
        
        # Initialize database
        db_manager = DatabaseManager(db_path=args.database, echo=False)
        
        with db_manager.get_session() as session:
            # Collect variants
            variants = collect_unknowns_and_low_confidence(
                session,
                days=args.days,
                confidence_threshold=args.confidence_threshold,
                min_frequency=args.min_frequency
            )
            
            if not variants:
                logger.warning("No variants found matching criteria")
                print("\n⚠️  No variants found matching criteria. Review queue not generated.")
                return
            
            # Cluster variants (unless disabled)
            if not args.no_clustering:
                variants = cluster_variants(variants)
            
            # Create Excel review queue
            create_excel_review_queue(variants, args.output, session)
        
        # Print summary
        print("\n" + "="*60)
        print("REVIEW QUEUE GENERATION SUMMARY")
        print("="*60)
        print(f"Variants collected:  {len(variants)}")
        print(f"Output file:         {args.output}")
        print(f"Lookback period:     {args.days} days")
        print(f"Confidence cutoff:   {args.confidence_threshold}")
        print("="*60)
        print(f"\n✅ Review queue generated successfully!")
        print(f"\nNext steps:")
        print(f"  1. Open {args.output} in Excel")
        print(f"  2. Fill in 'chosen_match' column using dropdowns")
        print(f"  3. Select 'validation_confidence' (HIGH/MEDIUM/LOW)")
        print(f"  4. Run: python scripts/12_validate_and_learn.py --review-queue {args.output}")
        
        logger.info("Review queue generation completed successfully!")
        
    except Exception as e:
        logger.error(f"Review queue generation failed: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()

"""
Batch matching CLI tool for chemical name resolution.

Processes lab EDD files (Excel/CSV) and generates:
- matched_results.xlsx: All results with confidence scores
- review_queue.xlsx: Flagged/unknown cases for human validation

Usage:
    python scripts/11_match_batch.py --input data/raw/lab_edds/sample_edd.xlsx --output reports/matched_results.xlsx
    python scripts/11_match_batch.py --input data.csv --output matched.xlsx --review-queue review.xlsx --confidence-threshold 0.75
"""

import argparse
import hashlib
import json
import logging
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy.orm import Session
from tqdm import tqdm

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from src.database.connection import DatabaseManager
from src.database import crud_new
from src.matching.resolution_engine import ResolutionEngine
from src.matching.match_result import ResolutionResult

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/batch_matching.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def detect_analyte_column(df: pd.DataFrame) -> Optional[str]:
    """
    Auto-detect which column contains analyte/parameter names.
    
    Looks for common column names used in Ontario lab EDDs.
    
    Args:
        df: DataFrame to analyze
    
    Returns:
        Column name or None if not detected
    """
    # Common column names in Ontario lab EDDs
    candidate_names = [
        'parameter', 'analyte', 'chemical', 'compound',
        'parameter name', 'analyte name', 'chemical name',
        'test', 'test name', 'analysis', 'method',
        'cas', 'cas_rn', 'substance'
    ]
    
    # Check exact matches (case-insensitive)
    columns_lower = {col.lower(): col for col in df.columns}
    for candidate in candidate_names:
        if candidate in columns_lower:
            logger.info(f"Auto-detected analyte column: '{columns_lower[candidate]}'")
            return columns_lower[candidate]
    
    # Check partial matches
    for col in df.columns:
        col_lower = col.lower()
        if any(name in col_lower for name in ['parameter', 'analyte', 'chemical']):
            logger.info(f"Auto-detected analyte column (partial match): '{col}'")
            return col
    
    logger.warning("Could not auto-detect analyte column")
    return None


def load_input_file(file_path: str, analyte_column: Optional[str] = None) -> Tuple[pd.DataFrame, str]:
    """
    Load input file (Excel or CSV) and extract analyte names.
    
    Args:
        file_path: Path to input file
        analyte_column: Column name containing analytes (auto-detect if None)
    
    Returns:
        Tuple of (DataFrame, detected_column_name)
    
    Raises:
        ValueError: If file format is unsupported or analyte column not found
    """
    file_path = Path(file_path)
    
    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")
    
    logger.info(f"Loading input file: {file_path}")
    
    # Load file based on extension
    if file_path.suffix.lower() in ['.xlsx', '.xls']:
        df = pd.read_excel(file_path)
    elif file_path.suffix.lower() == '.csv':
        df = pd.read_csv(file_path)
    else:
        raise ValueError(f"Unsupported file format: {file_path.suffix}")
    
    logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns")
    
    # Detect or validate analyte column
    if analyte_column is None:
        analyte_column = detect_analyte_column(df)
        if analyte_column is None:
            raise ValueError(
                f"Could not auto-detect analyte column. Please specify with --column.\n"
                f"Available columns: {', '.join(df.columns)}"
            )
    else:
        if analyte_column not in df.columns:
            raise ValueError(
                f"Column '{analyte_column}' not found in file.\n"
                f"Available columns: {', '.join(df.columns)}"
            )
    
    return df, analyte_column


def compute_corpus_hash(session: Session) -> str:
    """
    Compute hash of current corpus state for versioning.
    
    Args:
        session: Database session
    
    Returns:
        SHA256 hash of corpus
    """
    # Simple hash based on synonym count and timestamp
    from src.database.models import Synonym
    from sqlalchemy import select, func
    
    count = session.execute(select(func.count(Synonym.id))).scalar_one()
    hash_input = f"corpus_v1_{count}_{datetime.utcnow().date()}"
    return hashlib.sha256(hash_input.encode()).hexdigest()[:16]


def process_batch(
    analyte_names: List[str],
    session: Session,
    confidence_threshold: float = 0.75
) -> List[Dict]:
    """
    Process batch of analyte names through resolution engine.
    
    Args:
        analyte_names: List of analyte name strings
        session: Database session
        confidence_threshold: Minimum confidence threshold
    
    Returns:
        List of result dictionaries
    """
    engine = ResolutionEngine(session)
    results = []
    
    corpus_hash = compute_corpus_hash(session)
    model_hash = "fuzzy_v1"  # Placeholder for model versioning
    
    logger.info(f"Processing {len(analyte_names)} analytes with threshold={confidence_threshold}")
    
    for analyte_name in tqdm(analyte_names, desc="Matching analytes"):
        # Skip empty/null values
        if pd.isna(analyte_name) or str(analyte_name).strip() == '':
            results.append({
                'original_name': analyte_name,
                'matched_analyte_id': None,
                'matched_preferred_name': None,
                'cas_number': None,
                'confidence': 0.0,
                'match_method': 'skipped',
                'review_flag': False,
                'top_3_candidates': '[]',
                'resolution_time_ms': 0.0
            })
            continue
        
        try:
            # Resolve the name
            result = engine.resolve(str(analyte_name), confidence_threshold)
            
            # Extract top 3 candidates
            top_3 = []
            for candidate in result.all_candidates[:3]:
                top_3.append({
                    'analyte_id': candidate.analyte_id,
                    'name': candidate.preferred_name,
                    'confidence': round(candidate.confidence, 3),
                    'method': candidate.method
                })
            
            # Get analyte details if matched
            cas_number = None
            if result.best_match:
                analyte = crud_new.get_analyte_by_id(session, result.best_match.analyte_id)
                cas_number = analyte.cas_number if analyte else None
            
            # Determine review flag
            review_flag = result.requires_review or result.confidence_band == "UNKNOWN"
            
            # Build result record
            result_record = {
                'original_name': analyte_name,
                'matched_analyte_id': result.best_match.analyte_id if result.best_match else None,
                'matched_preferred_name': result.best_match.preferred_name if result.best_match else None,
                'cas_number': cas_number,
                'confidence': round(result.confidence, 3),
                'match_method': result.best_match.method if result.best_match else 'unknown',
                'review_flag': review_flag,
                'top_3_candidates': json.dumps(top_3),
                'resolution_time_ms': round(result.resolution_time_ms, 2)
            }
            
            results.append(result_record)
            
            # Log decision to database
            crud_new.log_match_decision(
                session=session,
                input_text=str(analyte_name),
                matched_analyte_id=result.best_match.analyte_id if result.best_match else None,
                match_method=result.best_match.method if result.best_match else 'unknown',
                confidence_score=result.confidence,
                top_k_candidates=top_3,
                signals_used=result.signals_used,
                corpus_snapshot_hash=corpus_hash,
                model_hash=model_hash,
                disagreement_flag=result.disagreement_flag
            )
            
        except Exception as e:
            logger.error(f"Error processing '{analyte_name}': {e}")
            results.append({
                'original_name': analyte_name,
                'matched_analyte_id': None,
                'matched_preferred_name': None,
                'cas_number': None,
                'confidence': 0.0,
                'match_method': 'error',
                'review_flag': True,
                'top_3_candidates': '[]',
                'resolution_time_ms': 0.0
            })
    
    # Commit all logged decisions
    session.commit()
    
    return results


def generate_summary_report(results: List[Dict]) -> str:
    """
    Generate summary statistics report.
    
    Args:
        results: List of match result dictionaries
    
    Returns:
        Formatted summary string
    """
    total = len(results)
    
    # Count by method
    exact_count = sum(1 for r in results if r['match_method'] in ['exact', 'cas_extracted'])
    fuzzy_count = sum(1 for r in results if r['match_method'] == 'fuzzy')
    semantic_count = sum(1 for r in results if r['match_method'] == 'semantic')
    unknown_count = sum(1 for r in results if r['match_method'] in ['unknown', 'error', 'skipped'])
    
    # Count by review flag
    review_flagged = sum(1 for r in results if r['review_flag'])
    
    # Average confidence (for matched only)
    confidences = [r['confidence'] for r in results if r['confidence'] > 0]
    avg_confidence = sum(confidences) / len(confidences) if confidences else 0.0
    
    # Processing time
    total_time_ms = sum(r['resolution_time_ms'] for r in results)
    
    # Build report
    report = f"""
╔══════════════════════════════════════════════════════════════╗
║           BATCH MATCHING SUMMARY REPORT                      ║
╚══════════════════════════════════════════════════════════════╝

Total Analytes Processed: {total:,}

┌─ Match Results ─────────────────────────────────────────────┐
│ Auto-accepted (exact):     {exact_count:6,} ({exact_count/total*100:5.1f}%)
│ Auto-accepted (fuzzy):     {fuzzy_count:6,} ({fuzzy_count/total*100:5.1f}%)
│ Semantic matches:          {semantic_count:6,} ({semantic_count/total*100:5.1f}%)
│ Review flagged:            {review_flagged:6,} ({review_flagged/total*100:5.1f}%)
│ Unknown/Error:             {unknown_count:6,} ({unknown_count/total*100:5.1f}%)
└─────────────────────────────────────────────────────────────┘

┌─ Corpus Maturity Indicators ────────────────────────────────┐
│ Exact match rate:          {exact_count/total*100:5.1f}%
│ Unknown rate:              {unknown_count/total*100:5.1f}%
│ Average confidence:        {avg_confidence:.3f}
│ Processing time:           {total_time_ms/1000:.1f} seconds
│ Throughput:                {total/(total_time_ms/1000):.0f} analytes/sec
└─────────────────────────────────────────────────────────────┘

Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
    
    return report


def write_excel_output(results: List[Dict], output_path: str, sheet_name: str = "Matched Results"):
    """
    Write results to Excel file with formatting.
    
    Args:
        results: List of result dictionaries
        output_path: Path to output Excel file
        sheet_name: Name of worksheet
    """
    df = pd.DataFrame(results)
    
    # Create Excel writer
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get worksheet for formatting
        ws = writer.sheets[sheet_name]
        
        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Color-code review flags
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red_fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
        
        review_col_idx = df.columns.get_loc('review_flag') + 1
        confidence_col_idx = df.columns.get_loc('confidence') + 1
        
        for row_idx, row in enumerate(df.itertuples(), start=2):
            if row.review_flag:
                for col_idx in range(1, len(df.columns) + 1):
                    if row.confidence == 0:
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill
                    else:
                        ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
    
    logger.info(f"Excel output written to: {output_path}")


def main():
    """Main entry point for batch matching CLI."""
    parser = argparse.ArgumentParser(
        description="Batch chemical name matching tool",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage with auto-detection
  python scripts/11_match_batch.py --input data/raw/lab_edds/sample.xlsx --output results.xlsx
  
  # Specify column and threshold
  python scripts/11_match_batch.py --input data.csv --column "Parameter" --output matched.xlsx --confidence-threshold 0.80
  
  # Generate separate review queue
  python scripts/11_match_batch.py --input data.xlsx --output matched.xlsx --review-queue review.xlsx
        """
    )
    
    parser.add_argument('--input', '-i', required=True, help='Input file (Excel or CSV)')
    parser.add_argument('--output', '-o', required=True, help='Output file for matched results')
    parser.add_argument('--review-queue', '-r', help='Output file for review queue (flagged cases only)')
    parser.add_argument('--column', '-c', help='Column name containing analyte names (auto-detect if not specified)')
    parser.add_argument('--confidence-threshold', '-t', type=float, default=0.75, 
                        help='Confidence threshold for auto-acceptance (default: 0.75)')
    parser.add_argument('--database', '-d', help='Path to database file (default: data/reg153_matcher.db)')
    
    args = parser.parse_args()
    
    # Validate threshold
    if not 0.0 <= args.confidence_threshold <= 1.0:
        logger.error(f"Confidence threshold must be between 0 and 1, got {args.confidence_threshold}")
        sys.exit(1)
    
    try:
        # Load input file
        logger.info(f"Starting batch matching job")
        logger.info(f"Input: {args.input}")
        logger.info(f"Output: {args.output}")
        logger.info(f"Confidence threshold: {args.confidence_threshold}")
        
        df, analyte_column = load_input_file(args.input, args.column)
        
        # Extract unique analyte names
        analyte_names = df[analyte_column].dropna().unique().tolist()
        logger.info(f"Found {len(analyte_names)} unique analyte names in column '{analyte_column}'")
        
        # Initialize database
        db_manager = DatabaseManager(db_path=args.database, echo=False)
        
        with db_manager.get_session() as session:
            # Process batch
            start_time = time.time()
            results = process_batch(analyte_names, session, args.confidence_threshold)
            elapsed_time = time.time() - start_time
        
        # Generate summary report
        summary = generate_summary_report(results)
        print(summary)
        
        # Write main output
        write_excel_output(results, args.output)
        
        # Write review queue if specified
        if args.review_queue:
            review_cases = [r for r in results if r['review_flag']]
            if review_cases:
                write_excel_output(review_cases, args.review_queue, sheet_name="Review Queue")
                logger.info(f"Review queue written to: {args.review_queue} ({len(review_cases)} cases)")
            else:
                logger.info("No review cases to write")
        
        # Save summary to file
        summary_path = Path(args.output).parent / f"summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        with open(summary_path, 'w') as f:
            f.write(summary)
        logger.info(f"Summary saved to: {summary_path}")
        
        logger.info("Batch matching completed successfully!")
        
    except Exception as e:
        logger.error(f"Batch matching failed: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()

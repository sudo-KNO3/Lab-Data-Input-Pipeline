"""
Script 40: Export Matched Lab Results

Produces a formatted Excel deliverable from ingested + matched lab data.
Cross-references lab_results.db with reg153_matcher.db to enrich each
row with Reg 153 analyte metadata (CAS, group, table number).

Modes:
    --submission-id N       Export a single submission
    --all                   Export all validated submissions
    --since YYYY-MM-DD      Export submissions received since a date
    --output PATH           Output file path (default: reports/exports/)
    --pivot                 Pivot results by sample (samples as columns)
    --include-non-chemical  Include rows marked as not_chemical

Output tabs:
    Results         — One row per chemical result, enriched with Reg 153 metadata
    Summary         — Statistics: match rates, chemical groups, confidence
    Submission Info — File metadata, vendor, dates

Usage:
    python scripts/40_export_results.py --submission-id 3
    python scripts/40_export_results.py --all --output reports/exports/full_export.xlsx
    python scripts/40_export_results.py --since 2026-02-01 --pivot
"""

import argparse
import logging
import sqlite3
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

LOG_DIR = PROJECT_ROOT / "logs"
LOG_DIR.mkdir(parents=True, exist_ok=True)
EXPORT_DIR = PROJECT_ROOT / "reports" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

LAB_DB = PROJECT_ROOT / "data" / "lab_results.db"
MATCHER_DB = PROJECT_ROOT / "data" / "reg153_matcher.db"

# ── Styling constants ─────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
GREY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

SUMMARY_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUMMARY_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

logger = logging.getLogger("export_results")


# ═══════════════════════════════════════════════════════════════════════════════
#  Data Loading
# ═══════════════════════════════════════════════════════════════════════════════

def load_analyte_lookup() -> Dict[str, dict]:
    """Load analyte metadata from reg153_matcher.db into a lookup dict."""
    conn = sqlite3.connect(str(MATCHER_DB))
    rows = conn.execute("""
        SELECT a.analyte_id, a.preferred_name, a.cas_number, a.group_code,
               a.chemical_group, a.table_number, a.parent_analyte_id,
               p.preferred_name AS parent_name
        FROM analytes a
        LEFT JOIN analytes p ON a.parent_analyte_id = p.analyte_id
    """).fetchall()
    conn.close()

    lookup = {}
    for row in rows:
        lookup[row[0]] = {
            "preferred_name": row[1],
            "cas_number": row[2],
            "group_code": row[3],
            "chemical_group": row[4],
            "table_number": row[5],
            "parent_analyte_id": row[6],
            "parent_name": row[7],
        }
    return lookup


def get_submissions(
    submission_id: Optional[int] = None,
    export_all: bool = False,
    since: Optional[str] = None,
) -> List[dict]:
    """Fetch submission metadata from lab_results.db."""
    conn = sqlite3.connect(str(LAB_DB))
    conn.row_factory = sqlite3.Row

    if submission_id:
        rows = conn.execute(
            "SELECT * FROM lab_submissions WHERE submission_id = ?",
            (submission_id,),
        ).fetchall()
    elif since:
        rows = conn.execute(
            "SELECT * FROM lab_submissions WHERE received_date >= ? ORDER BY submission_id",
            (since,),
        ).fetchall()
    elif export_all:
        rows = conn.execute(
            "SELECT * FROM lab_submissions ORDER BY submission_id"
        ).fetchall()
    else:
        rows = []

    result = [dict(r) for r in rows]
    conn.close()
    return result


def get_results(
    submission_ids: List[int],
    include_non_chemical: bool = False,
) -> pd.DataFrame:
    """Fetch lab results for given submission IDs."""
    if not submission_ids:
        return pd.DataFrame()

    conn = sqlite3.connect(str(LAB_DB))
    placeholders = ",".join("?" * len(submission_ids))

    status_filter = ""
    if not include_non_chemical:
        status_filter = "AND r.validation_status != 'not_chemical'"

    query = f"""
        SELECT
            r.submission_id,
            s.original_filename,
            s.lab_vendor,
            s.received_date,
            r.row_number,
            r.chemical_raw,
            r.chemical_normalized,
            r.analyte_id,
            r.correct_analyte_id,
            r.match_method,
            r.match_confidence,
            r.sample_id,
            r.result_value,
            r.units,
            r.qualifier,
            r.detection_limit,
            r.lab_method,
            r.validation_status,
            r.human_override,
            r.medium
        FROM lab_results r
        JOIN lab_submissions s ON r.submission_id = s.submission_id
        WHERE r.submission_id IN ({placeholders})
        {status_filter}
        ORDER BY r.submission_id, r.row_number
    """
    df = pd.read_sql_query(query, conn, params=submission_ids)
    conn.close()
    return df


# ═══════════════════════════════════════════════════════════════════════════════
#  Data Enrichment
# ═══════════════════════════════════════════════════════════════════════════════

def enrich_results(df: pd.DataFrame, analyte_lookup: Dict[str, dict]) -> pd.DataFrame:
    """
    Enrich lab results with Reg 153 analyte metadata.
    Uses correct_analyte_id if human-overridden, otherwise analyte_id.
    """
    if df.empty:
        return df

    # Determine effective analyte ID (human correction takes precedence)
    df["effective_analyte_id"] = df.apply(
        lambda row: row["correct_analyte_id"]
        if pd.notna(row["correct_analyte_id"]) and row["correct_analyte_id"]
        else row["analyte_id"],
        axis=1,
    )

    # Enrich with analyte metadata
    df["reg153_name"] = df["effective_analyte_id"].map(
        lambda aid: analyte_lookup.get(aid, {}).get("preferred_name", "")
    )
    df["cas_number"] = df["effective_analyte_id"].map(
        lambda aid: analyte_lookup.get(aid, {}).get("cas_number", "")
    )
    df["chemical_group"] = df["effective_analyte_id"].map(
        lambda aid: analyte_lookup.get(aid, {}).get("chemical_group", "")
    )
    df["group_code"] = df["effective_analyte_id"].map(
        lambda aid: analyte_lookup.get(aid, {}).get("group_code", "")
    )
    df["table_number"] = df["effective_analyte_id"].map(
        lambda aid: analyte_lookup.get(aid, {}).get("table_number", "")
    )

    # Confidence band for display
    df["confidence_band"] = df["match_confidence"].apply(_confidence_label)

    return df


def _confidence_label(conf) -> str:
    if conf is None or pd.isna(conf):
        return "UNMATCHED"
    if conf >= 0.95:
        return "HIGH"
    if conf >= 0.75:
        return "MEDIUM"
    if conf > 0:
        return "LOW"
    return "UNMATCHED"


# ═══════════════════════════════════════════════════════════════════════════════
#  Excel Writer
# ═══════════════════════════════════════════════════════════════════════════════

# Column definitions for the Results sheet
RESULTS_COLUMNS = [
    ("Lab File", "original_filename"),
    ("Lab Vendor", "lab_vendor"),
    ("Sample ID", "sample_id"),
    ("Medium", "medium"),
    ("Lab Chemical Name", "chemical_raw"),
    ("Reg 153 Analyte", "reg153_name"),
    ("Analyte ID", "effective_analyte_id"),
    ("CAS Number", "cas_number"),
    ("Chemical Group", "chemical_group"),
    ("Reg 153 Table", "table_number"),
    ("Result", "result_value"),
    ("Units", "units"),
    ("Qualifier", "qualifier"),
    ("Detection Limit", "detection_limit"),
    ("Lab Method", "lab_method"),
    ("Match Method", "match_method"),
    ("Match Confidence", "match_confidence"),
    ("Confidence Band", "confidence_band"),
    ("Validation Status", "validation_status"),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  Convenience API (for watcher / programmatic use)
# ═══════════════════════════════════════════════════════════════════════════════

def export_submission(
    submission_id: int,
    output_path: str,
    lab_db_path: Optional[str] = None,
    matcher_db_path: Optional[str] = None,
) -> Path:
    """
    Export a single submission to a formatted Excel file.

    This is the public API for programmatic export (e.g. from the
    file watcher).  It handles all the data loading internally.
    """
    # Temporarily override module-level DB paths if provided
    global LAB_DB, MATCHER_DB
    orig_lab, orig_matcher = LAB_DB, MATCHER_DB
    try:
        if lab_db_path:
            LAB_DB = Path(lab_db_path)
        if matcher_db_path:
            MATCHER_DB = Path(matcher_db_path)

        submissions = get_submissions(submission_id=submission_id)
        if not submissions:
            raise ValueError(f"Submission #{submission_id} not found")

        sub_ids = [s["submission_id"] for s in submissions]
        df = get_results(sub_ids)

        if df.empty:
            raise ValueError(f"No results for submission #{submission_id}")

        analyte_lookup = load_analyte_lookup()
        df = enrich_results(df, analyte_lookup)

        return write_export(df, submissions, Path(output_path))
    finally:
        LAB_DB, MATCHER_DB = orig_lab, orig_matcher


def write_export(
    df: pd.DataFrame,
    submissions: List[dict],
    output_path: Path,
    pivot: bool = False,
) -> Path:
    """Write the formatted Excel export."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # ── Tab 1: Results ─────────────────────────────────────────────
        _write_results_sheet(df, writer, pivot)

        # ── Tab 2: Summary ─────────────────────────────────────────────
        _write_summary_sheet(df, submissions, writer)

        # ── Tab 3: Submission Info ─────────────────────────────────────
        _write_submissions_sheet(submissions, writer)

    logger.info(f"Exported to {output_path}")
    return output_path


def _write_results_sheet(df: pd.DataFrame, writer: pd.ExcelWriter, pivot: bool):
    """Write the main results tab."""
    if df.empty:
        pd.DataFrame({"Note": ["No results to export"]}).to_excel(
            writer, sheet_name="Results", index=False
        )
        return

    if pivot and "sample_id" in df.columns:
        _write_pivot_sheet(df, writer)
        return

    # Build export DataFrame with clean column names
    export_data = {}
    for display_name, col_key in RESULTS_COLUMNS:
        if col_key in df.columns:
            export_data[display_name] = df[col_key].values
        else:
            export_data[display_name] = ""
    export_df = pd.DataFrame(export_data)

    export_df.to_excel(writer, sheet_name="Results", index=False, startrow=0)
    ws = writer.sheets["Results"]

    # Style header
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # Row-level color coding by confidence band
    band_col = export_df.columns.get_loc("Confidence Band")
    for row_idx in range(2, len(export_df) + 2):
        band_val = ws.cell(row=row_idx, column=band_col + 1).value
        fill = _band_fill(band_val)
        for col_idx in range(1, len(export_df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Format confidence column as percentage
    conf_col = export_df.columns.get_loc("Match Confidence") + 1
    for row_idx in range(2, len(export_df) + 2):
        cell = ws.cell(row=row_idx, column=conf_col)
        if cell.value is not None and cell.value != "":
            try:
                cell.number_format = "0.0%"
            except (ValueError, TypeError):
                pass

    # Auto-width columns
    _auto_width(ws, max_width=45)

    # Freeze top row + first 3 columns
    ws.freeze_panes = "D2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _write_pivot_sheet(df: pd.DataFrame, writer: pd.ExcelWriter):
    """Write results pivoted: chemicals as rows, samples as columns."""
    # Only include rows with sample_id
    has_sample = df[df["sample_id"].notna() & (df["sample_id"] != "")]
    if has_sample.empty:
        pd.DataFrame({"Note": ["No sample-level data to pivot"]}).to_excel(
            writer, sheet_name="Results", index=False
        )
        return

    # Build display value: qualifier + result
    has_sample = has_sample.copy()
    has_sample["display_value"] = has_sample.apply(
        lambda r: f"{r['qualifier'] or ''}{r['result_value'] or ''}", axis=1
    )

    # Pivot: rows = chemical, columns = sample_id
    pivot_df = has_sample.pivot_table(
        index=["reg153_name", "chemical_group", "cas_number", "units"],
        columns="sample_id",
        values="display_value",
        aggfunc="first",
    )
    pivot_df = pivot_df.reset_index()
    pivot_df.columns.name = None

    # Rename index columns
    pivot_df = pivot_df.rename(columns={
        "reg153_name": "Reg 153 Analyte",
        "chemical_group": "Chemical Group",
        "cas_number": "CAS Number",
        "units": "Units",
    })

    pivot_df.to_excel(writer, sheet_name="Results", index=False)
    ws = writer.sheets["Results"]

    # Style header
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    _auto_width(ws, max_width=20)
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(
    df: pd.DataFrame, submissions: List[dict], writer: pd.ExcelWriter
):
    """Write summary statistics tab."""
    rows = []
    rows.append(("EXPORT SUMMARY", ""))
    rows.append(("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")))
    rows.append(("Submissions", len(submissions)))
    rows.append(("Total Results", len(df)))
    rows.append(("", ""))

    # Confidence breakdown
    rows.append(("MATCH CONFIDENCE", "Count"))
    for band in ["HIGH", "MEDIUM", "LOW", "UNMATCHED"]:
        count = len(df[df["confidence_band"] == band]) if not df.empty else 0
        rows.append((f"  {band}", count))
    rows.append(("", ""))

    # Match method breakdown
    rows.append(("MATCH METHOD", "Count"))
    if not df.empty:
        for method, count in df["match_method"].value_counts().items():
            rows.append((f"  {method}", count))
    rows.append(("", ""))

    # Chemical group breakdown
    rows.append(("CHEMICAL GROUP", "Count"))
    if not df.empty:
        for group, count in (
            df[df["chemical_group"] != ""]["chemical_group"]
            .value_counts()
            .items()
        ):
            rows.append((f"  {group}", count))
    rows.append(("", ""))

    # Vendor breakdown
    rows.append(("LAB VENDOR", "Submissions"))
    if not df.empty:
        for vendor, count in df["lab_vendor"].value_counts().items():
            rows.append((f"  {vendor}", count))

    summary_df = pd.DataFrame(rows, columns=["Metric", "Value"])
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    ws = writer.sheets["Summary"]

    # Style header
    for cell in ws[1]:
        cell.fill = SUMMARY_HEADER_FILL
        cell.font = SUMMARY_HEADER_FONT
        cell.alignment = HEADER_ALIGN

    # Bold section headers
    bold_font = Font(bold=True, size=11)
    for row_idx in range(2, len(summary_df) + 2):
        cell_a = ws.cell(row=row_idx, column=1)
        val = str(cell_a.value or "")
        if val and not val.startswith("  ") and val != "":
            cell_a.font = bold_font
            cell_b = ws.cell(row=row_idx, column=2)
            cell_b.font = bold_font

    _auto_width(ws, max_width=40)


def _write_submissions_sheet(submissions: List[dict], writer: pd.ExcelWriter):
    """Write submission metadata tab."""
    if not submissions:
        pd.DataFrame({"Note": ["No submissions"]}).to_excel(
            writer, sheet_name="Submissions", index=False
        )
        return

    display_cols = [
        ("Submission ID", "submission_id"),
        ("Original Filename", "original_filename"),
        ("Lab Vendor", "lab_vendor"),
        ("Received Date", "received_date"),
        ("Validation Status", "validation_status"),
        ("Extraction Accuracy", "extraction_accuracy"),
    ]
    data = {}
    for display_name, key in display_cols:
        data[display_name] = [s.get(key, "") for s in submissions]

    sub_df = pd.DataFrame(data)
    sub_df.to_excel(writer, sheet_name="Submissions", index=False)
    ws = writer.sheets["Submissions"]

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    _auto_width(ws, max_width=60)


# ═══════════════════════════════════════════════════════════════════════════════
#  Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def _band_fill(band: str) -> PatternFill:
    if band == "HIGH":
        return GREEN_FILL
    elif band == "MEDIUM":
        return YELLOW_FILL
    elif band == "LOW":
        return RED_FILL
    elif band == "UNMATCHED":
        return RED_FILL
    return GREY_FILL


def _auto_width(ws, max_width: int = 50):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_width)


# ═══════════════════════════════════════════════════════════════════════════════
#  CLI
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="Export matched lab results to formatted Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        "--submission-id", type=int,
        help="Export a single submission by ID",
    )
    group.add_argument(
        "--all", action="store_true", dest="export_all",
        help="Export all validated submissions",
    )
    group.add_argument(
        "--since", type=str,
        help="Export submissions received since YYYY-MM-DD",
    )

    parser.add_argument(
        "--output", "-o", type=str, default=None,
        help="Output file path (default: reports/exports/<auto>.xlsx)",
    )
    parser.add_argument(
        "--pivot", action="store_true",
        help="Pivot results by sample (samples as columns)",
    )
    parser.add_argument(
        "--include-non-chemical", action="store_true",
        help="Include rows marked as not_chemical (footers, calculated params)",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="Verbose logging",
    )

    args = parser.parse_args()

    # Setup logging
    level = logging.DEBUG if args.verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        handlers=[logging.StreamHandler()],
    )

    # Validate DB files exist
    if not LAB_DB.exists():
        logger.error(f"Lab results database not found: {LAB_DB}")
        sys.exit(1)
    if not MATCHER_DB.exists():
        logger.error(f"Matcher database not found: {MATCHER_DB}")
        sys.exit(1)

    # Load analyte lookup
    analyte_lookup = load_analyte_lookup()
    logger.info(f"Loaded {len(analyte_lookup)} analytes from matcher DB")

    # Fetch submissions
    submissions = get_submissions(
        submission_id=args.submission_id,
        export_all=args.export_all,
        since=args.since,
    )
    if not submissions:
        logger.warning("No submissions found matching criteria.")
        sys.exit(0)

    sub_ids = [s["submission_id"] for s in submissions]
    logger.info(f"Exporting {len(submissions)} submission(s): {sub_ids}")

    # Fetch and enrich results
    df = get_results(sub_ids, include_non_chemical=args.include_non_chemical)
    logger.info(f"Loaded {len(df)} result rows")

    df = enrich_results(df, analyte_lookup)

    # Determine output path
    if args.output:
        output_path = Path(args.output)
    else:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if args.submission_id:
            name = f"submission_{args.submission_id}_{ts}.xlsx"
        elif args.since:
            name = f"results_since_{args.since}_{ts}.xlsx"
        else:
            name = f"all_results_{ts}.xlsx"
        output_path = EXPORT_DIR / name

    # Write export
    out = write_export(df, submissions, output_path, pivot=args.pivot)

    # Print summary
    print(f"\n{'='*60}")
    print(f"EXPORT COMPLETE")
    print(f"{'='*60}")
    print(f"  File:         {out}")
    print(f"  Submissions:  {len(submissions)}")
    print(f"  Results:      {len(df)}")
    if not df.empty:
        high = len(df[df["confidence_band"] == "HIGH"])
        med = len(df[df["confidence_band"] == "MEDIUM"])
        low = len(df[df["confidence_band"] == "LOW"])
        unm = len(df[df["confidence_band"] == "UNMATCHED"])
        print(f"  Confidence:   {high} high, {med} medium, {low} low, {unm} unmatched")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()

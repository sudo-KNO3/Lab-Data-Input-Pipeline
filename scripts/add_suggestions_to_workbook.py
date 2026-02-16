"""
Alternative validation approach: Add top suggestions as comments/extra columns.
This avoids dropdown dependency if Excel version doesn't support them well.
"""
import sqlite3
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from src.database.connection import DatabaseManager
from src.normalization.text_normalizer import TextNormalizer
from src.matching.resolution_engine import ResolutionEngine


def add_suggestions_to_workbook(submission_id: int):
    """Add top match suggestions as cell comments."""
    
    # Load existing workbook
    wb_path = Path(f"reports/validation/validation_{submission_id}_FIXED.xlsx")
    if not wb_path.exists():
        print(f"✗ File not found: {wb_path}")
        return
    
    wb = load_workbook(str(wb_path))
    ws = wb["🔬 Chemical Review"]
    
    # Get lab results
    lab_conn = sqlite3.connect("data/lab_results.db")
    results = lab_conn.execute("""
        SELECT result_id, chemical_raw, chemical_normalized, match_confidence
        FROM lab_results
        WHERE submission_id = ?
        ORDER BY row_number
    """, (submission_id,)).fetchall()
    lab_conn.close()
    
    # Get top suggestions for low-confidence matches
    db = DatabaseManager()
    normalizer = TextNormalizer()
    
    with db.get_session() as session:
        resolver = ResolutionEngine(session, normalizer)
        
        for idx, (result_id, chem_raw, chem_norm, conf) in enumerate(results, start=2):
            # Skip high-confidence matches
            if conf and conf >= 0.95:
                continue
            
            # Get top 5 suggestions
            result = resolver.resolve(chem_norm, confidence_threshold=0.50)
            
            if result.all_candidates:
                # Create suggestion text
                suggestions = []
                for i, match in enumerate(result.all_candidates[:5], 1):
                    suggestions.append(
                        f"{i}. {match.preferred_name} ({match.analyte_id}) - {match.confidence:.1%}"
                    )
                
                suggestion_text = "Top suggestions:\n" + "\n".join(suggestions)
                
                # Add as comment to "Corrected Match" cell
                cell = ws[f"G{idx}"]
                cell.comment = Comment(suggestion_text, "System")
                
                print(f"   Row {idx}: Added {len(suggestions)} suggestions for '{chem_raw}'")
    
    # Save with comments
    output_path = wb_path.parent / f"validation_{submission_id}_WITH_SUGGESTIONS.xlsx"
    wb.save(str(output_path))
    
    print(f"\n✓ Created: {output_path}")
    print(f"   Hover over cells in 'Corrected Match' column to see suggestions!")
    
    return output_path


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--submission-id', type=int, required=True)
    args = parser.parse_args()
    
    path = add_suggestions_to_workbook(args.submission_id)
    
    if path:
        print(f"\n   Open: {path}")

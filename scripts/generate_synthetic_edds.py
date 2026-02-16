"""
Generate synthetic lab EDD (Electronic Data Deliverable) files for testing.

Creates realistic Excel files mimicking Ontario environmental lab reports from:
- ALS Environmental
- SGS Canada
- Bureau Veritas

Includes realistic Ontario lab naming variants, typos, and formatting patterns.
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


# Base analytes with multiple naming variants per lab
ANALYTE_VARIANTS = {
    'Benzene': {
        'canonical': 'Benzene',
        'cas': '71-43-2',
        'als': ['Benzene', 'Benezene', 'Benzene '],  # Includes typo and spacing
        'sgs': ['Benzene', 'BENZENE'],
        'bv': ['Benzene'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 8260'
    },
    'Toluene': {
        'canonical': 'Toluene',
        'cas': '108-88-3',
        'als': ['Toluene', 'Toluenne', 'Toluene '],
        'sgs': ['Toluene', 'TOLUENE'],
        'bv': ['Toluene'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 8260'
    },
    'Trichloroethylene': {
        'canonical': 'Trichloroethylene',
        'cas': '79-01-6',
        'als': ['Trichloroethylene', 'TCE', 'Trichloroethene'],
        'sgs': ['Trichloroethylene', 'TCE'],
        'bv': ['Trichloroethylene (TCE)'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 8260'
    },
    '1,1,1-Trichloroethane': {
        'canonical': '1,1,1-Trichloroethane',
        'cas': '71-55-6',
        'als': ['1,1,1-Trichloroethane', '1,1,1-TCA', '1, 1, 1-TCA', '1,1,1 TCA'],
        'sgs': ['1,1,1-Trichloroethane', '1,1,1-TCA'],
        'bv': ['1,1,1-Trichloroethane'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 8260'
    },
    '1,4-Dioxane': {
        'canonical': '1,4-Dioxane',
        'cas': '123-91-1',
        'als': ['1,4-Dioxane', '1,4 Diox', '1,4-Dioxane '],
        'sgs': ['1,4-Dioxane', 'Dioxane, 1,4-'],
        'bv': ['1,4-Dioxane'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 8260'
    },
    'Benzo(a)pyrene': {
        'canonical': 'Benzo(a)pyrene',
        'cas': '50-32-8',
        'als': ['Benzo(a)pyrene', 'B(a)P', 'BaP', 'Benzo[a]pyrene'],
        'sgs': ['Benzo(a)pyrene', 'B(a)P'],
        'bv': ['Benzo(a)pyrene'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 8270'
    },
    'Naphthalene': {
        'canonical': 'Naphthalene',
        'cas': '91-20-3',
        'als': ['Naphthalene', 'Naphthalene '],
        'sgs': ['Naphthalene', 'NAPHTHALENE'],
        'bv': ['Naphthalene'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 8270'
    },
    'PHC F2': {
        'canonical': 'Petroleum Hydrocarbons F2 (>C10-C16)',
        'cas': '',
        'als': ['PHC F2', 'PHC F2 (C10-C16)', 'Petroleum Hydrocarbons F2'],
        'sgs': ['PHC F2', 'F2 Petroleum Hydrocarbons'],
        'bv': ['PHC F2 (C10-C16)'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'CCME F2'
    },
    'PHC F3': {
        'canonical': 'Petroleum Hydrocarbons F3 (>C16-C34)',
        'cas': '',
        'als': ['PHC F3', 'PHC F3 (C16-C34)', 'Petroleum Hydrocarbons F3'],
        'sgs': ['PHC F3', 'F3 Petroleum Hydrocarbons'],
        'bv': ['PHC F3 (C16-C34)'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'CCME F3'
    },
    'Chromium': {
        'canonical': 'Chromium',
        'cas': '7440-47-3',
        'als': ['Chromium', 'Cr'],
        'sgs': ['Chromium, Total', 'Chromium (Total)'],
        'bv': ['Chromium, Total'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 6010'
    },
    'Lead': {
        'canonical': 'Lead',
        'cas': '7439-92-1',
        'als': ['Lead', 'Pb'],
        'sgs': ['Lead, Total', 'Lead (Total)'],
        'bv': ['Lead, Total'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 6010'
    },
    'Arsenic': {
        'canonical': 'Arsenic',
        'cas': '7440-38-2',
        'als': ['Arsenic', 'As'],
        'sgs': ['Arsenic, Total', 'Arsenic (Total)'],
        'bv': ['Arsenic, Total'],
        'units': {'Soil': 'mg/kg', 'Groundwater': 'µg/L'},
        'method': 'EPA 6010'
    },
}


def generate_sample_ids(num_samples, sample_type='mixed'):
    """Generate realistic sample IDs."""
    sample_ids = []
    
    for i in range(num_samples):
        if sample_type == 'mixed':
            if random.random() < 0.6:  # 60% soil samples
                sample_ids.append(f"S{random.randint(1, 99):03d}-{random.randint(1, 20):03d}")
            else:  # 40% groundwater
                sample_ids.append(f"MW-{random.randint(1, 30):02d}-GW")
        elif sample_type == 'soil':
            sample_ids.append(f"S{random.randint(1, 99):03d}-{random.randint(1, 20):03d}")
        else:
            sample_ids.append(f"MW-{random.randint(1, 30):02d}-GW")
    
    return sample_ids


def generate_collection_dates(num_dates):
    """Generate random collection dates in the past 6 months."""
    end_date = datetime.now()
    start_date = end_date - timedelta(days=180)
    
    dates = []
    for _ in range(num_dates):
        random_days = random.randint(0, 180)
        date = start_date + timedelta(days=random_days)
        dates.append(date.strftime('%Y-%m-%d'))
    
    return dates


def generate_result_value(matrix, analyte_info, is_detect=None):
    """Generate realistic result value with lognormal distribution."""
    if is_detect is None:
        is_detect = random.random() < 0.6  # 60% detect rate
    
    # Base detection limits
    if matrix == 'Soil':
        mdl = random.uniform(0.01, 0.1)
        rl = mdl * random.uniform(2, 5)
    else:  # Groundwater
        mdl = random.uniform(0.1, 1.0)
        rl = mdl * random.uniform(2, 5)
    
    if is_detect:
        # Detected - lognormal distribution above RL
        mu = np.log(rl * random.uniform(3, 50))
        sigma = random.uniform(0.5, 1.5)
        value = np.exp(np.random.normal(mu, sigma))
        qualifier = random.choice(['', '', '', 'J', 'B'])  # Most no qualifier
    else:
        # Non-detect
        value = rl
        qualifier = 'U'
    
    return round(value, 3), round(mdl, 3), round(rl, 3), qualifier


def generate_als_edd(output_path, num_samples=20, num_analytes_per_sample=None):
    """Generate ALS Environmental style EDD."""
    
    if num_analytes_per_sample is None:
        num_analytes_per_sample = random.randint(8, 15)
    
    print(f"Generating ALS EDD with {num_samples} samples...")
    
    # Create Excel writer
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Cover sheet
    cover_data = {
        'Item': [
            'Laboratory', 'Report Date', 'Project Name', 'Project Number',
            'Client', 'Report Number', 'Number of Samples'
        ],
        'Value': [
            'ALS Environmental', datetime.now().strftime('%Y-%m-%d'),
            'Site Assessment - Main Street', 'PROJ-2025-1234',
            'ABC Environmental Consulting', f'ALS-{random.randint(100000, 999999)}',
            str(num_samples)
        ]
    }
    df_cover = pd.DataFrame(cover_data)
    df_cover.to_excel(writer, sheet_name='Cover', index=False)
    
    # Generate sample data
    sample_ids = generate_sample_ids(num_samples)
    client_ids = [f"Client-{sid}" for sid in sample_ids]
    collection_dates = generate_collection_dates(num_samples)
    matrices = ['Soil' if 'S' in sid else 'Groundwater' for sid in sample_ids]
    
    # Sample Summary sheet
    summary_data = {
        'Lab_Sample_ID': sample_ids,
        'Client_Sample_ID': client_ids,
        'Collection_Date': collection_dates,
        'Matrix': matrices,
        'Tests_Requested': ['VOC, SVOC, PHC, Metals'] * num_samples
    }
    df_summary = pd.DataFrame(summary_data)
    df_summary.to_excel(writer, sheet_name='Sample Summary', index=False)
    
    # Results sheet - expand each sample to multiple analytes
    results_data = []
    
    for i in range(num_samples):
        # Select random analytes for this sample
        selected_analytes = random.sample(list(ANALYTE_VARIANTS.keys()), 
                                         min(num_analytes_per_sample, len(ANALYTE_VARIANTS)))
        
        for analyte_name in selected_analytes:
            analyte_info = ANALYTE_VARIANTS[analyte_name]
            
            # Randomly select variant name for ALS
            parameter_name = random.choice(analyte_info['als'])
            
            # Generate result
            value, mdl, rl, qualifier = generate_result_value(matrices[i], analyte_info)
            
            results_data.append({
                'Lab_Sample_ID': sample_ids[i],
                'Client_Sample_ID': client_ids[i],
                'Collection_Date': collection_dates[i],
                'Matrix': matrices[i],
                'Analysis_Method': analyte_info['method'],
                'Parameter': parameter_name,
                'CAS_Number': analyte_info['cas'] if random.random() < 0.9 else '',  # 10% missing CAS
                'Result_Value': value if qualifier != 'U' else f"<{rl}",
                'Units': analyte_info['units'][matrices[i]],
                'MDL': mdl,
                'RL': rl,
                'Qualifier_Flags': qualifier
            })
    
    df_results = pd.DataFrame(results_data)
    df_results.to_excel(writer, sheet_name='Results', index=False)
    
    # QA/QC sheet (Excel doesn't allow "/" in sheet names)
    qaqc_data = {
        'QC_Sample_Type': ['Method Blank', 'Laboratory Control Sample', 'Matrix Spike', 'Duplicate'],
        'Analysis_Method': ['EPA 8260', 'EPA 8270', 'EPA 6010', 'EPA 8260'],
        'Parameter': ['Benzene', 'Naphthalene', 'Lead, Total', 'TCE'],
        'Result': ['<0.001', '95%', '102%', '5.2% RPD'],
        'Acceptance_Criteria': ['<RL', '70-130%', '75-125%', '<20% RPD'],
        'Status': ['Pass', 'Pass', 'Pass', 'Pass']
    }
    df_qaqc = pd.DataFrame(qaqc_data)
    df_qaqc.to_excel(writer, sheet_name='QAQC', index=False)
    
    writer.close()
    
    print(f"  Created {len(results_data)} result rows")
    print(f"  Saved to: {output_path}")


def generate_sgs_edd(output_path, num_samples=15, num_analytes_per_sample=None):
    """Generate SGS Canada style EDD."""
    
    if num_analytes_per_sample is None:
        num_analytes_per_sample = random.randint(8, 15)
    
    print(f"Generating SGS EDD with {num_samples} samples...")
    
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Cover sheet
    cover_data = {
        'Field': [
            'Laboratory', 'Report Date', 'Project ID', 'Client Name',
            'SGS Report ID', 'Total Samples', 'Revision'
        ],
        'Information': [
            'SGS Canada Inc. - Environmental', datetime.now().strftime('%Y-%m-%d'),
            f'SGS-{random.randint(10000, 99999)}', 'XYZ Remediation Ltd.',
            f'SGS-ENV-{random.randint(100000, 999999)}', str(num_samples), 'Rev 0'
        ]
    }
    df_cover = pd.DataFrame(cover_data)
    df_cover.to_excel(writer, sheet_name='Report Information', index=False)
    
    # Generate sample data
    sample_ids = [f"SGS{random.randint(100000, 999999)}" for _ in range(num_samples)]
    client_ids = [f"Field-{random.randint(1, 999):03d}" for _ in range(num_samples)]
    collection_dates = generate_collection_dates(num_samples)
    matrices = random.choices(['Soil', 'Groundwater'], k=num_samples)
    
    # Sample Information sheet
    sample_info_data = {
        'SGS_Sample_ID': sample_ids,
        'Field_Sample_ID': client_ids,
        'Sample_Date': collection_dates,
        'Sample_Type': matrices,
        'Received_Date': [datetime.now().strftime('%Y-%m-%d')] * num_samples,
        'Analysis_Suite': ['Full Organics + Metals'] * num_samples
    }
    df_sample_info = pd.DataFrame(sample_info_data)
    df_sample_info.to_excel(writer, sheet_name='Sample Information', index=False)
    
    # Analytical Results sheet - SGS style with "Analyte" instead of "Parameter"
    results_data = []
    
    for i in range(num_samples):
        selected_analytes = random.sample(list(ANALYTE_VARIANTS.keys()), 
                                         min(num_analytes_per_sample, len(ANALYTE_VARIANTS)))
        
        for analyte_name in selected_analytes:
            analyte_info = ANALYTE_VARIANTS[analyte_name]
            
            # Use SGS variant names
            analyte_display = random.choice(analyte_info['sgs'])
            
            value, mdl, rl, qualifier = generate_result_value(matrices[i], analyte_info)
            
            results_data.append({
                'SGS_Sample_ID': sample_ids[i],
                'Field_Sample_ID': client_ids[i],
                'Sample_Date': collection_dates[i],
                'Matrix': matrices[i],
                'Method': analyte_info['method'],
                'Analyte': analyte_display,  # SGS uses "Analyte" not "Parameter"
                'CAS_RN': analyte_info['cas'],
                'Result': value if qualifier != 'U' else f"<{rl}",
                'Unit': analyte_info['units'][matrices[i]],
                'RL': rl,  # SGS typically shows RL not MDL
                'Flag': qualifier
            })
    
    df_results = pd.DataFrame(results_data)
    df_results.to_excel(writer, sheet_name='Analytical Results', index=False)
    
    # Quality Control
    qc_data = {
        'QC_Type': ['Blank', 'LCS', 'MS/MSD', 'Replicate'],
        'Method': ['EPA 8260', 'EPA 8270', 'EPA 6010', 'EPA 8260'],
        'Analyte': ['Benzene', 'Naphthalene', 'Chromium, Total', 'Trichloroethylene'],
        'Recovery_%': ['<RL', '98', '105/103', 'n/a'],
        'RPD_%': ['n/a', 'n/a', '1.9', '4.2'],
        'Accept': ['Y', 'Y', 'Y', 'Y']
    }
    df_qc = pd.DataFrame(qc_data)
    df_qc.to_excel(writer, sheet_name='Quality Control', index=False)
    
    writer.close()
    
    print(f"  Created {len(results_data)} result rows")
    print(f"  Saved to: {output_path}")


def generate_bureauveritas_edd(output_path, num_samples=18, num_analytes_per_sample=None):
    """Generate Bureau Veritas style EDD with nested headers."""
    
    if num_analytes_per_sample is None:
        num_analytes_per_sample = random.randint(8, 15)
    
    print(f"Generating Bureau Veritas EDD with {num_samples} samples...")
    
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Cover sheet
    cover_data = {
        'Report Information': [
            'Laboratory Name', 'Issue Date', 'Project Reference', 'Client',
            'BV Report Number', 'LIMS Batch ID', 'Sample Count'
        ],
        '': [
            'Bureau Veritas Laboratories', datetime.now().strftime('%Y-%m-%d'),
            f'BV-PROJ-{random.randint(1000, 9999)}', 'Environmental Solutions Inc.',
            f'BV-{random.randint(200000, 299999)}', f'LIMS-{random.randint(10000, 99999)}',
            str(num_samples)
        ]
    }
    df_cover = pd.DataFrame(cover_data)
    df_cover.to_excel(writer, sheet_name='Report Cover', index=False)
    
    # Generate sample data with BV style IDs
    sample_ids = [f"BV{datetime.now().year}{random.randint(10000, 99999)}" for _ in range(num_samples)]
    lims_ids = [f"LIMS-{random.randint(100000, 999999)}" for _ in range(num_samples)]
    client_ids = [f"Sample-{random.randint(1, 200):03d}" for _ in range(num_samples)]
    collection_dates = generate_collection_dates(num_samples)
    matrices = random.choices(['Soil', 'Groundwater'], k=num_samples)
    
    # Sample Registry
    registry_data = {
        'BV_Lab_ID': sample_ids,
        'LIMS_ID': lims_ids,
        'Client_ID': client_ids,
        'Collection_Date': collection_dates,
        'Matrix_Type': matrices,
        'Test_Package': ['ORG-METALS-COMBO'] * num_samples,
        'Receipt_Condition': ['Intact'] * num_samples
    }
    df_registry = pd.DataFrame(registry_data)
    df_registry.to_excel(writer, sheet_name='Sample Registry', index=False)
    
    # Data Results - Bureau Veritas style with unique spacing
    results_data = []
    
    for i in range(num_samples):
        selected_analytes = random.sample(list(ANALYTE_VARIANTS.keys()), 
                                         min(num_analytes_per_sample, len(ANALYTE_VARIANTS)))
        
        for analyte_name in selected_analytes:
            analyte_info = ANALYTE_VARIANTS[analyte_name]
            
            # Use BV variant names
            analyte_display = random.choice(analyte_info['bv'])
            
            value, mdl, rl, qualifier = generate_result_value(matrices[i], analyte_info)
            
            results_data.append({
                'BV_Lab_ID': sample_ids[i],
                'LIMS_ID': lims_ids[i],
                'Client_ID': client_ids[i],
                'Date_Collected': collection_dates[i],
                'Matrix': matrices[i],
                'Test_Method': analyte_info['method'],
                'Analyte_Name': analyte_display,
                'CAS_Registry': analyte_info['cas'] if analyte_info['cas'] else 'N/A',
                'Measured_Value': value if qualifier != 'U' else rl,
                'Result_Units': analyte_info['units'][matrices[i]],
                'Detection_Limit': mdl,
                'Reporting_Limit': rl,
                'Data_Qualifier': qualifier,
                'ND_Flag': 'Y' if qualifier == 'U' else 'N'
            })
    
    df_results = pd.DataFrame(results_data)
    df_results.to_excel(writer, sheet_name='Data Results', index=False)
    
    # Close writer before modifying with openpyxl
    writer.close()
    
    # Apply nested header formatting (multi-row headers)
    wb = openpyxl.load_workbook(output_path)
    ws = wb['Data Results']
    
    # Add a header row above
    ws.insert_rows(1)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Group columns with nested headers
    ws['A1'] = 'Sample Identification'
    ws['D1'] = 'Sample Identification'
    ws['E1'] = 'Sample Info'
    ws['F1'] = 'Analysis'
    ws['I1'] = 'Results'
    ws['M1'] = 'Flags'
    
    for cell in ['A1', 'D1', 'E1', 'F1', 'I1', 'M1']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center')
    
    wb.save(output_path)
    
    print(f"  Created {len(results_data)} result rows")
    print(f"  Saved to: {output_path}")


def main():
    """Generate all synthetic EDDs."""
    
    # Base path
    base_path = Path(r"n:\Central\Staff\KG\Kiefer's Coding Corner\Reg 153 chemical matcher")
    output_dir = base_path / "data" / "raw" / "lab_edds"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    print("="*70)
    print("GENERATING SYNTHETIC LAB EDD FILES")
    print("="*70)
    
    # Generate each lab format
    generate_als_edd(
        output_dir / "ALS_example_report.xlsx",
        num_samples=25,
        num_analytes_per_sample=12
    )
    
    print()
    generate_sgs_edd(
        output_dir / "SGS_example_report.xlsx",
        num_samples=20,
        num_analytes_per_sample=10
    )
    
    print()
    generate_bureauveritas_edd(
        output_dir / "BureauVeritas_example_report.xlsx",
        num_samples=22,
        num_analytes_per_sample=11
    )
    
    print()
    print("="*70)
    print("GENERATION COMPLETE")
    print("="*70)
    print(f"\nFiles created in: {output_dir}")
    print("\nNext steps:")
    print("  1. Run verification script to validate structure")
    print("  2. Use these files to test normalization and matching pipelines")
    print("  3. Review ontario_variants_known.csv for ground truth patterns")


if __name__ == "__main__":
    main()
